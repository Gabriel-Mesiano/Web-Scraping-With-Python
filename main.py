import time
import openpyxl as op
import tabulate as tb
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

# Adiciona opção de rodar em Background
option = webdriver.ChromeOptions()
option.add_argument('--headless=none')
# option.add_argument('--start-maximized')

# Inicializa o webdriver do Chrome
driver = webdriver.Chrome(options=option)

# Escolha algo para pesquisar no google
search = 'Kabum'

# Abre o endereço no Chrome
driver.get(f'https://www.google.com.br/search?q={search}')
try:
    # Clica no primeiro link
    driver.find_element(By.XPATH, '//*[@id="rso"]/div[1]/div/div/div/div/div/div/div/div[1]/div/span/a/h3').click()
    # Procura pela barra de search, escreve e da enter
    time.sleep(1)
    searchBars = driver.find_elements(By.TAG_NAME, 'input')
    for x in searchBars:
        if x.get_attribute('id') and ('search' in x.get_attribute('id') or 'busca' in x.get_attribute('id')) :
            x.click()
            x.send_keys("ps5")
            x.send_keys(Keys.ENTER)
            break
    
    
    # Cria uma tabela com os preços para o site Kabum
    productCards = driver.find_elements(By.CLASS_NAME, 'productCard')
    listNames = []
    listPrice = []
    col = ["Nome", "Preço"]
    for x in productCards:
        listNames.append(x.find_element(By.CLASS_NAME, 'nameCard').text)
        listPrice.append(x.find_element(By.CLASS_NAME, 'priceCard').text)

    planilha = op.Workbook()
    cell = planilha['Sheet']
    cell.title = 'Primeira planilha'
    cell['A1'] = 'Nome'
    cell['B1'] = 'Preço'
    index = 2
    for name, price in zip(listNames, listPrice):
        cell.cell(column=1,row=index, value=name)
        cell.cell(column=2,row=index, value=price)
        index += 1
    planilha.save('primeira_planilha.xlsx')
finally:
    # Fecha o Chrome
    time.sleep(5)
    driver.quit()